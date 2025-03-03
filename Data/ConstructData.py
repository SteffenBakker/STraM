# -*- coding: utf-8 -*-
"""
Created on Mon Nov  8 13:31:04 2021

@author: ingvi
Adapted by Ruben and Steffen
"""

"Example data"

import os
import sys
import pickle

standalone_testing = False
if standalone_testing:
    #os.chdir('M:/Documents/GitHub/AIM_Norwegian_Freight_Model') #uncomment this for stand-alone testing of this fille
    #os.chdir('C:\\Users\\steffejb\\OneDrive - NTNU\\Work\\GitHub\\STRAM_ntnu_development') # STEFFEN
    #os.chdir('C:\\Github\\STRAM_ntnu_development') # RUBEN NTNU
    os.chdir('C:\\Users\\Ruben\\Github\\STRAM_ntnu_development') # RUBEN EUR
    sys.path.insert(0, '') #make sure the modules are found in the new working directory

from Data.settings import *
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string, get_column_letter
import itertools
import numpy as np
import matplotlib.pyplot as plt

from math import cos, asin, sqrt, pi
import networkx as nx
from itertools import islice
pd.set_option("display.max_rows", None, "display.max_columns", None)
import numpy as np
from Data.BassDiffusion import BassDiffusion 
from Data.sigmoid import sigmoid
from path_generation import path_generation

# from FreightTransportModel.Utils import plot_all_graphs  #FreightTransportModel.

from functools import wraps
import time

def timeit(func):
    @wraps(func)
    def timeit_wrapper(*args, **kwargs):
        start_time = time.perf_counter()
        result = func(*args, **kwargs)
        end_time = time.perf_counter()
        total_time = end_time - start_time
        print(f'Function {func.__name__} Took {total_time:.4f} seconds')
        return result
    return timeit_wrapper

# Translate scenario_tree to the name of the corresponding excel sheet
def get_scen_sheet_name(scenario_tree):
    sheet_name_scenarios = ""
    if scenario_tree == 'AllScen':
        sheet_name_scenarios = 'scenarios_base' 
    elif scenario_tree == '4Scen':
        sheet_name_scenarios = 'four_scenarios' 
    elif scenario_tree == '9Scen':
        sheet_name_scenarios = 'nine_scenarios' 
    elif scenario_tree == "FuelScen":
        sheet_name_scenarios = 'fuel_scenarios'
    elif scenario_tree == "FuelDetScen":
        sheet_name_scenarios = 'fuel_scenario_det'
    return sheet_name_scenarios

#Class containing information about all scenarios
#Information in this class can be used to activate a scenario in a TransportSets object, meaning that the corresponding parameter values are changed
class ScenarioInformation():
    def __init__(self, sh_name='fuel_scenarios'): 

        #TODO: REVAMP HOW WE DO SCENARIOS. PROBABLY DELETE A BUNCH OF CODE BELOW
              
        # first, read modes and fuel groups
        mode_sets = pd.read_excel(r'Data/sets.xlsx', sheet_name = "modes")
        fuel_sets = pd.read_excel(r'Data/sets.xlsx', sheet_name = "fuels")

        M_MODES = []       # all modes
        F_FUEL = []            # all fuels
        FM_FUEL = {}           # all fuels per mode
        NEW_MF_LIST = []       # all new fuels                        
        FG_FUEL_GROUPS = []    # all fuel groups
        F_TO_FG = {}           # fuel to fuel group

        for index, row in mode_sets.iterrows():
            cur_mode = row["Mode"]
            M_MODES.append(cur_mode)

        # initialize FM_FUEL
        for m in M_MODES:
            FM_FUEL[m] = []

        for index, row in fuel_sets.iterrows():
            cur_mode = row["Mode"]
            cur_fuel = row["Fuel"]
            cur_fuel_group = row["Fuel group"]
            if(cur_fuel not in F_FUEL):
                F_FUEL.append(cur_fuel)
            FM_FUEL[cur_mode].append(cur_fuel)
            if(cur_fuel_group) not in FG_FUEL_GROUPS:
                FG_FUEL_GROUPS.append(cur_fuel_group)
            F_TO_FG[cur_fuel] = cur_fuel_group

        # store relevant data in scenario object        
        self.fuel_group_names = FG_FUEL_GROUPS
        self.fuel_groups = {} #dict from fuel group names to (m,f) combinations
        # initialize 
        for fg in self.fuel_group_names:
            self.fuel_groups[fg] = []
        # fill with correct elements
        for m in M_MODES:
            for f in FM_FUEL[m]:
                cur_fuel_group = F_TO_FG[f]
                self.fuel_groups[cur_fuel_group].append((m,f))
        
        self.mf_to_fg = {} #translate (m,f) to fuel group name
        for fg in self.fuel_groups:
            for (m,f) in self.fuel_groups[fg]:
                self.mf_to_fg[(m,f)] = fg

    
        # read and process scenario data
        scenario_data = pd.read_excel(r'Data/'+"scenarios.xlsx", sheet_name=sh_name) 
        self.num_scenarios = len(scenario_data)
        self.scenario_names = ["scen_" + str(i).zfill(len(str(self.num_scenarios))) for i in range(self.num_scenarios)] #initialize as scen_00, scen_01, scen_02, etc.
        self.probabilities = [1.0/self.num_scenarios] * self.num_scenarios #initialize as equal probabilities
        
        variability_data = pd.read_excel(r'Data/'+"scenarios.xlsx", sheet_name="variability")
        cost_factor_data = pd.read_excel(r'Data/'+"scenarios.xlsx", sheet_name="cost_factor")


        self.variability = None
        for index, row in variability_data.iterrows():
           if index == 0:
               self.variability = row["Variability"]
        
        self.cost_factor = {}
        for index, row in cost_factor_data.iterrows():
            #fg = row['Fuel Group']
            s = row['Scenario']
            t = row['Time Period']
            m = row['Mode']
            p = row['Product Group']
            f = row['Fuel']
            self.cost_factor[(s,t,m,p,f)]= row['Cost Factor']
        
        #To get a value for each year
        years = sorted(list(set([k[1] for k in self.cost_factor.keys()])))
        new_dict = {}
        for s,_,m,p,f in self.cost_factor.keys():
            for k in range(len(years)-1):
                delta = (self.cost_factor[(s,years[k+1],m,p,f)]- self.cost_factor[(s,years[k],m,p,f)])/ (years[k+1] - years[k])
                for y in range(years[k]+1, years[k+1]):
                    new_dict[(s,y,m,p,f)] = self.cost_factor[(s,years[k],m,p,f)] + delta * (y-years[k])
        
        self.cost_factor.update(new_dict)


        self.fg_maturity_path_name = [{}] * self.num_scenarios  
        self.fg_fuel_cost_path_name = [{}] * self.num_scenarios  
        #self.fg_cost_factor = [{}] * self.num_scenarios #this is already explicitly done in the cost_factor sheet
        for index, row in scenario_data.iterrows():
            if "Name" in scenario_data:
                self.scenario_names[index] = row["Name"] #update scenario names if available
            if "Probability" in scenario_data:
                self.probabilities[index] = row["Probability"] #update probabilities if available
            for fg in self.fuel_group_names:
                #new_maturity_entry = {fg : row[fg]}
                #self.fg_maturity_path_name[index] = dict(self.fg_maturity_path_name[index], **new_maturity_entry) #add new entry to existing dict (trick from internet)
                new_cost_path_entry = {fg: row[fg]}
                self.fg_fuel_cost_path_name[index] = dict(self.fg_fuel_cost_path_name[index], **new_cost_path_entry) #add new entry to existing dict (trick from internet)
                
                # new_cost_factor = None
                # if row[fg] == "B":
                #     new_cost_factor = 1.0
                # elif row[fg] ==  "O":
                #     new_cost_factor = 1.0 - self.variability
                # elif row[fg] == "P":
                #     new_cost_factor = 1.0 + self.variability
                # new_cost_entry = {fg : new_cost_factor} #new entry for the dictionary fg_cost_factor[index]
                # self.fg_cost_factor[index] = dict(self.fg_cost_factor[index], **new_cost_entry) #add new entry to existing dict (trick from internet)
        
        

        #make dicts for scenario name to nr and vice versa
        self.scen_name_to_nr = {}
        self.scen_nr_to_name = {}
        for i in range(len(self.scenario_names)):
            self.scen_name_to_nr[self.scenario_names[i]] = i
            self.scen_nr_to_name[i] = self.scenario_names[i]
        
        # NB: the following is now depreciated as we directly read this from excel
        # self.mode_fuel_cost_factor = [] #list of dictionaries (per scenario) from (m,f) pair to transport cost factor (relative to base cost)
        # for s in range(self.num_scenarios):
        #     self.mode_fuel_cost_factor.append({})
        #     for fg in self.fuel_group_names:
        #         for mf in self.fuel_groups[fg]:
        #             self.mode_fuel_cost_factor[s][mf] = self.fg_cost_factor[s][fg]



#Class containing all relevant data
#Note: also contains the scenario information (in self.scenario_information)
#One scenario can be activated (indicated by self.active_scenario_nr) by the procedure update_scenario_dependent_parameters(self,scenario_nr)
#Activating a scenario means that all relevant parameters are changed to their scenario values
class TransportSets():

    def __init__(self,sheet_name_scenarios='fuel_scenarios',co2_fee="base", TIMES_data=None):# or (self) 
        
        self.single_time_period = None #only solve last time period -> remove all operational constraints for the other periods

        #read/construct scenario information
        self.active_scenario_name = "benchmark" #no scenario has been activated; all data is from benchmark setting
        self.scenario_information = ScenarioInformation(sheet_name_scenarios)
        self.scenario_information_EV = ScenarioInformation('EV_scenario') 

        self.risk_information = None

        #read/construct data                
        self.construct_pyomo_data(co2_fee, TIMES_data)
        self.combined_sets(TIMES_data)
        
        

    def construct_pyomo_data(self,co2_fee, TIMES_data=None):

        print("Reading and constructing data")

        self.scaling_factor = SCALING_FACTOR #10E-5
        self.scaling_factor_monetary = SCALING_FACTOR_MONETARY
        self.scaling_factor_weight = SCALING_FACTOR_WEIGHT
        self.scaling_factor_emissions = SCALING_FACTOR_EMISSIONS

        self.precision_digits = NUM_DIGITS_PRECISION #this is necessary, otherwise C_TRANSP dissapears!

        self.S_SCENARIOS_ALL = self.scenario_information.scenario_names
        self.S_SCENARIOS = self.S_SCENARIOS_ALL
        
        self.construct_sets(*[pd.read_excel(r'Data/sets.xlsx',
                                            sheet_name=s_name) 
                              for s_name in ("modes",
                                             "fuels",
                                             "products",
                                             "time_periods")])
        
        
        
        self.construct_network(*[pd.read_excel(r'Data/SPATIAL/spatial_data.xlsx',
                                               sheet_name=s_name) 
                              for s_name in ("zones_STRAM",
                                             "distances_STRAM")])
        
        self.construct_ODD(pd.read_csv(r'Data/SPATIAL/demand.csv'), TIMES_data)
        
        cost_input = load_workbook(r'Data/cost_calculator.xlsx',
                                   data_only=True)["Parameter Input"]
        
        self.construct_vehicles(cost_input)

        self.construct_emission_transfer(pd.read_excel(r'Data/cost_calculator.xlsx',
                                                sheet_name="Transfer costs"),
                                         cost_input,
                                         co2_fee)
        
        if TIMES_data is not None:
            conv_file = pd.read_excel(r'linking/Spatial/spatial_conversion_STRAM_TIMES2.xlsx',
                                      sheet_name='edges').fillna(0)
        else:
            conv_file = None
        
        self.construct_costs(pd.read_excel(r'Data/cost_calculator.xlsx',
                              sheet_name="Parameter Input"), conv_file, TIMES_data) #TODO : changer feuille si usage link
        
        self.construct_time_value(*[pd.read_excel(r'Data/time_value.xlsx',
                                                sheet_name=s_name) 
                                  for s_name in ("Output",
                                                 "Speeds")])
        
        self.construct_investments(*[pd.read_excel(r'Data/capacities_and_investments.xlsx',
                                                   sheet_name=s_name) 
                              for s_name in ("node_capacities",
                                             "edge_capacities")])
        
        self.construct_charging_edges(pd.read_excel(r'Data/capacities_and_investments.xlsx',
                                                    sheet_name='charging_data'))
        
        inp1 = [pd.read_excel(r'Data/technological_maturity_readiness.xlsx',
                              sheet_name=s_name) 
                              for s_name in ("technological_readiness_bass",
                                             "phase_out_fuels")]
        inp2 = [pd.read_excel(r'Data/init_mode_fuel_mix.xlsx',
                              sheet_name=s_name) 
                              for s_name in ("init_fuel_mix",
                                             "init_mode_mix")]
        self.construct_tech_readiness(*inp1, *inp2)
        

        self.construct_path_generation(r'Data/SPATIAL/generated_paths_'\
                                       +str(NUM_MODE_PATHS)+'_modes.pkl')
        
        self.construct_param2()
        
        
    @timeit
    def construct_sets(self, mode_sets, fuel_sets, product_sets, time_period_sets):

        # Modes and fuels
        
        self.M_MODES = []       # all modes
        self.M_MODES_CAP = []   # capacitated modes

        self.LIFETIME = {}      # lifetime of vehicle on each mode
        
        for index, row in mode_sets.iterrows():
            cur_mode = row["Mode"]
            cur_capacitated = row["Capacitated"]
            cur_lifetime = row["Lifetime"]
            self.M_MODES.append(cur_mode)
            if(cur_capacitated == "Yes"):
                self.M_MODES_CAP.append(cur_mode)
            self.LIFETIME[cur_mode] = cur_lifetime
        
        self.F_FUEL = []            # all fuels
        self.FM_FUEL = {}           # all fuels per mode
        self.NEW_MF_LIST = []       # all new fuels                        
        self.FG_FUEL_GROUPS = []    # all fuel groups
        self.F_TO_FG = {}           # fuel to fuel group

        # initialize FM_FUEL
        for m in self.M_MODES:
            self.FM_FUEL[m] = []

        for index, row in fuel_sets.iterrows():
            cur_mode = row["Mode"]
            cur_fuel = row["Fuel"]
            cur_fuel_group = row["Fuel group"]
            cur_novelty = row["Novelty"]
            if(cur_fuel not in self.F_FUEL):
                self.F_FUEL.append(cur_fuel)
            self.FM_FUEL[cur_mode].append(cur_fuel)
            if(cur_fuel_group) not in self.FG_FUEL_GROUPS:
                self.FG_FUEL_GROUPS.append(cur_fuel_group)
            self.F_TO_FG[cur_fuel] = cur_fuel_group
            if(cur_novelty == "New"):
                self.NEW_MF_LIST.append((cur_mode, cur_fuel))
        
        # list of new fuels
        self.NEW_F_LIST = set([e[1] for e in self.NEW_MF_LIST])

        # Products
    
        self.P_PRODUCTS = []
        self.PC_PRODUCT_CLASSES = []
        self.PC_TO_P = {}
        self.P_TO_PC = {}

        for index, row in product_sets.iterrows():
            cur_prod = row["Product group"]
            cur_prod_class = row["Product class"]
            self.P_PRODUCTS.append(cur_prod)
            if(cur_prod_class not in self.PC_PRODUCT_CLASSES):
                self.PC_PRODUCT_CLASSES.append(cur_prod_class)
                self.PC_TO_P[cur_prod_class] = [cur_prod]
            else:
                self.PC_TO_P[cur_prod_class].append(cur_prod)
            self.P_TO_PC[cur_prod] = cur_prod_class


        # Time periods

        self.T_TIME_PERIODS = []    # all time periods (years) with decisions
        
        self.T_TIME_FIRST_STAGE_BASE = []   # first-stage periods
        self.T_TIME_SECOND_STAGE_BASE = []  # second-stage periods

        for index, row in time_period_sets.iterrows():
            cur_year = row["Year"]
            cur_stage = row["Stage"]
            self.T_TIME_PERIODS.append(cur_year)
            if(cur_stage == 1):
                self.T_TIME_FIRST_STAGE_BASE.append(cur_year)
            else:
                self.T_TIME_SECOND_STAGE_BASE.append(cur_year)
            

        # previous period
        self.T_MIN1 = {self.T_TIME_PERIODS[tt]:self.T_TIME_PERIODS[tt-1] for tt in range(1,len(self.T_TIME_PERIODS))} 

        #we have to switch between solving only first time period, and all time periods. (to initialize the transport shares and emissions)
        self.T_TIME_PERIODS_ALL = self.T_TIME_PERIODS           # all time periods
        self.T_TIME_PERIODS_INIT = [self.T_TIME_PERIODS[0]]     # only first period (for initialization)

        # time periods for reading demand data
        self.T_TIME_PERIODS_PWC = [2018, 2020, 2025, 2030, 2040, 2050]  # this should still be HARDCODED

    @timeit
    def construct_network(self, zone_data, distances):

        #NODES
        self.N_NODES = []
        self.SEA_NODES = []
        self.ROAD_NODES = []
        self.RAIL_NODES = []
        self.N_ABROAD = []
        self.centroid_to_nr = {}
        self.centroid_to_zone = {}
        self.zone_nr_to_centroid = {}
        self.zone_to_centroid = {}

        for index, row in zone_data.iterrows():
            node = row["centroid_name"]
            self.N_NODES.append(node) 
            self.centroid_to_nr[node] = row["zone_nr"]
            self.zone_nr_to_centroid[row["zone_nr"]] = row["centroid_name"]
            self.zone_to_centroid[row["zone_name"]] = row["centroid_name"]
            self.centroid_to_zone[node] = row["zone_name"]
            if row["abroad"] == 1:
                self.N_ABROAD.append(node)
            if row["sea"] == 1:
                self.SEA_NODES.append(node)
            if row["road"] == 1:
                self.ROAD_NODES.append(node)
            if row["rail"] == 1:
                self.RAIL_NODES.append(node)


        self.N_NODES_NORWAY = list(set(self.N_NODES) - set(self.N_ABROAD))

        self.NM_NODES = {m:None for m in self.M_MODES}
        self.NM_NODES["Road"] = self.ROAD_NODES
        self.NM_NODES["Sea"] = self.SEA_NODES
        self.NM_NODES["Rail"] = self.RAIL_NODES

        self.M_MODES_CAP = ["Rail", "Sea"] #this should be done in capacity excel sheet
        #self.N_NODES_CAP_NORWAY = {"Rail": self.RAIL_NODES_NORWAY,"Sea": self.SEA_NODES_NORWAY} #this should be done in capacity excel sheet

        self.N_ZONE_NR = dict(zip(zone_data.centroid_name, zone_data.zone_nr))
        self.N_ZONE_NAME = dict(zip(zone_data.centroid_name, zone_data.zone_name))
        self.N_LATITUDE = dict(zip(zone_data.centroid_name, zone_data.latitude))
        self.N_LONGITUDE = dict(zip(zone_data.centroid_name, zone_data.longitude))
        self.N_LATITUDE_PLOT = dict(zip(zone_data.centroid_name, zone_data.lat_plot))
        self.N_LONGITUDE_PLOT = dict(zip(zone_data.centroid_name, zone_data.long_plot))
        self.N_COORD_PLOT = {zone_data.centroid_name[i]: (zone_data.long_plot[i], zone_data.lat_plot[i]) for i in range(len(zone_data.centroid_name))}
        self.N_COORD_OFFSETS = {zone_data.centroid_name[i]: (zone_data.offset_x[i], zone_data.offset_y[i]) for i in range(len(zone_data.centroid_name))}

        #EDGES and DISTANCES
        
        self.E_EDGES = []
        self.A_ARCS = []
        
        distances_dict = {}
        for index, row in distances.iterrows():
            i = self.zone_nr_to_centroid[row["From"]]
            j = self.zone_nr_to_centroid[row["To"]]
            distances_dict[(i,j,row["Mode"],int(row["Route"]))] = row["DistanceKM"]   

        self.DISTANCE = {}
        self.AVG_DISTANCE = {}
        for (i,j,m,r),value in distances_dict.items():
            a1 = (i,j,m,r)
            a2 = (j,i,m,r)
            self.E_EDGES.append(a1)
            self.A_ARCS.append(a1)
            self.A_ARCS.append(a2)
            self.DISTANCE[a1] = value
            self.DISTANCE[a2] = value
            self.AVG_DISTANCE[a1] = value
            self.AVG_DISTANCE[a2] = value

        self.AE_ARCS = {e:[] for e in self.E_EDGES}
        self.AM_ARCS = {m:[] for m in self.M_MODES}
        for (i,j,m,r) in self.E_EDGES:
            a1 = (i,j,m,r)
            a2 = (j,i,m,r)
            self.AE_ARCS[a1].append(a1)
            self.AE_ARCS[a1].append(a2)
            self.AM_ARCS[m].append(a1)
            self.AM_ARCS[m].append(a2)

        self.ANM_ARCS_IN = {(n,m):[] for n in self.N_NODES for m in self.M_MODES}
        self.ANM_ARCS_OUT = {(n,m):[] for n in self.N_NODES for m in self.M_MODES}
        for (i,j,m,r) in self.A_ARCS:
            a = (i,j,m,r)
            self.ANM_ARCS_IN[(j,m)].append(a)
            self.ANM_ARCS_OUT[(i,m)].append(a)

        # -----------------------
        # ------- Other--------
        # -----------------------

        if NO_DRY_BULK:
            self.P_PRODUCTS.remove("Dry bulk") #Dry bulk   HARDCODING
        if NO_LIQUID_BULK:
            self.P_PRODUCTS.remove("Liquid bulk") #Wet bulk

    @timeit
    def construct_ODD(self, pwc_aggr, TIMES_data):
        ####################################
        ### ORIGIN, DESTINATION AND DEMAND #
        ####################################


        #Start with filtering of demand data. What to include and what not.
        D_DEMAND_ALL = {}  #5330 entries, after cutting off the small elements, only 4105 entries

        #then read the pwc_aggr data
        for index, row in pwc_aggr.iterrows(): #WE ONLY TAKE DEMAND BETWEEN COUNTIES! SO, THIS IS OUR DEFINITION OF LONG-DISTANCE TRANSPORT
            from_zone = row['from_aggr_zone']
            to_zone = row['to_aggr_zone']
            from_node = self.zone_to_centroid[from_zone]
            to_node = self.zone_to_centroid[to_zone]
            product = row['product_group']
            if product in self.P_PRODUCTS:
                D_DEMAND_ALL[(from_node, to_node,product ,int(row['year']))] = round(float(row['amount_tons']),0)
            else:
                if product != 'Liquid bulk':
                    product
    

        demands = pd.Series(D_DEMAND_ALL.values())         
        
        if False: #Remove demand -> No, then we only end up with containers and dry bulk
            # DO THIS ANALYSIS AS TON/KM?, opposed to as in TONNES? should not matter too much. distances are somewhat similar
            print('describing original demand data')
            #print(demands.describe())   #huge spread -> remove the very small stuff. Even demand of 5E-1
            #demands.plot.hist(by=None, bins=1000)
            #demands.hist(cumulative=True, density=1, bins=100)
            
            total_base_demand=round(demands.sum(),0)  #'1.339356e+09' TONNES
            cutoff = demands.quantile(0.12) #HARDCODED, requires some analyses. Do not remove too much. These are actual things that need to be transported. And not too much complexity because of it
            print('cutoff (in Tonnes):', cutoff)
            demands2 = demands[demands > cutoff]  #and (demands < demands.quantile(0.9)
            reduced_base_demand = round(demands2.sum(),0)  #'1.338888e+09' TONNES
            print('percentage demand removed: ',(total_base_demand-reduced_base_demand)/total_base_demand*100,'%')  
            
            # demands2.plot.hist(by=None, bins=1000)
            # demands2.hist(cumulative=True, density=1, bins=100)
            
            # print("{:e}".format(round(cutoff,0)))    # '3.306000e+03'
            # print("{:e}".format(round(demands.max(),0)))           # '2.739037e+07'
        
            #D_DEMAND_CUT = D_DEMAND_ALL #
            D_DEMAND_CUT = {key:value for key,value in D_DEMAND_ALL.items() if value > cutoff}  #(o,d,p,t)
            
            #print('D_DEMAND_CUT:')
            #print(pd.Series(list(D_DEMAND_CUT.values())).describe())

            print(len(D_DEMAND_ALL))  #9432
            print(len(D_DEMAND_CUT))  #8299

        #ODP
        self.OD_PAIRS = {p: [] for p in self.P_PRODUCTS}
        for (o,d,p,t), value in D_DEMAND_ALL.items():
            if ((o,d) not in self.OD_PAIRS[p]):
                self.OD_PAIRS[p].append((o,d))
        self.ODP = []
        self.OD_PAIRS_ALL = set()
        for p in self.P_PRODUCTS:
            for (o, d) in self.OD_PAIRS[p]:
                self.OD_PAIRS_ALL.add((o, d))
                self.ODP.append((o, d, p))
        self.OD_PAIRS_ALL = list(self.OD_PAIRS_ALL)
        
        self.D_DEMAND = {(o,d,p,t):0 for t in self.T_TIME_PERIODS_PWC for (o,d,p) in self.ODP}        
        for (o,d,p,t), value in D_DEMAND_ALL.items():
            self.D_DEMAND[(o,d,p,t)] = round(value / self.scaling_factor_weight,self.precision_digits)
        
        # if TIMES_data != None:
        #     for (o,d,p,t), val in TIMES_data.trade_TIMES_odpt.items():
        #         self.D_DEMAND[(o,d,p,t)] += 0 # TODO val
        
        
        
        
        #print('D_DEMAND:')
        #print(pd.Series(list(self.D_DEMAND.values())).describe())

        #self.D_DEMAND = {key:round(value,self.precision_digits) for (key,value) in self.D_DEMAND.items()}
        #self.D_DEMAND = {key:value for key,value in self.D_DEMAND.items() if value > 0}


        D_DEMAND_AGGR = {(p,t):0 for t in self.T_TIME_PERIODS_PWC for p in self.P_PRODUCTS}
        D_DEMAND_AGGR_BASE_YEAR = {p:0 for p in self.P_PRODUCTS}
        for (o,d,p,t),value in self.D_DEMAND.items():
            D_DEMAND_AGGR[(p,t)] += value
            if t == self.T_TIME_PERIODS_PWC[0]:
                D_DEMAND_AGGR_BASE_YEAR[p] += value

        DEMAND_PRODUCT_BASE_YEAR=pd.DataFrame([[key, value] for key, value in D_DEMAND_AGGR_BASE_YEAR.items()],columns=['product','tot_weight'])
        DEMAND_PRODUCT_BASE_YEAR["tot_weight"] = round(DEMAND_PRODUCT_BASE_YEAR["tot_weight"]/10**6*SCALING_FACTOR_WEIGHT,2) # in MTonnes 
        print(f"Transport demand in the base year {self.T_TIME_PERIODS_PWC[0]} (in MTonnes):")
        print(DEMAND_PRODUCT_BASE_YEAR)

    @timeit
    def construct_vehicles(self, ws): #, data_only=True
        
        columns_MP = ["Year",
                      "Capex costumization",
                       "Residual value (costumization)",
                       "Opex fix (admin, crew, insurance)",
                       "Tonnage",
                       "Market failure (av. utilization)",
                       "Annual Mileage",
                       "Lifetime costumization",
                       "Opex var (mode-fees)"]
        
        columns_MF = ["Year",
                      "Capex vehicle",
                        "Residual value (vehicle)",
                        "Lifetime vehicle",
                        "Opex maintenance & repair",
                        "Fuel Economy",
                        "Fuel Cost",
                        "Payload loss (fuel)",
                        "WACC",
                        "Emission cost",
                        "Carbon permits",
                        "CO2 content per mass/kWh",
                        "Energy density/Efficiency",
                        "CO2 content per kWh",
                        "CO2 emission per km",
                        "CO2 emission cost per km"]
        
        self.VEHICLES_DATA_MP = {(m,p):pd.DataFrame(columns=columns_MP).set_index('Year') for m in self.M_MODES for p in self.PC_PRODUCT_CLASSES}
        self.VEHICLES_DATA_MF = {(m,f):pd.DataFrame(columns=columns_MF).set_index('Year') for m in self.M_MODES for f in self.FM_FUEL[m]}
        
        
        taxes = {}
        
        for row in range(1, ws.max_row+1):
            
            if ws.cell(row, 2).value == "TAX":
                taxes[(ws.cell(row, 3).value, ws.cell(row, 4).value)] = float(ws.cell(row, 5).value)
            
            if ws.cell(row, 1).value is not None and type(ws.cell(row, 1).value) == str :
                if "Road" in ws.cell(row, 1).value:
                    mode = "Road"
                elif "Railway" in ws.cell(row, 1).value:
                    mode = "Rail"
                elif "Sea" in ws.cell(row, 1).value:
                    mode = "Sea"
            
            if ws.cell(row, 1).value is not None and type(ws.cell(row, 1).value) == str and ws.cell(row, 1).value.split(": ")[-1] in self.PC_PRODUCT_CLASSES:
                product = ws.cell(row, 1).value.split(": ")[-1]
                t_start = int(ws.cell(row, 2).value)
                t_end = int(ws.cell(row, 3).value)
                for t in range(t_start, t_end+1):
                    values = [sigmoid(t,
                                        float(ws.cell(row+subrow, 2).value),
                                        float(ws.cell(row+subrow, 3).value),
                                        t_start, t_end,
                                        float(ws.cell(row+subrow, 5).value),
                                        float(ws.cell(row+subrow, 6).value),
                                        float(ws.cell(row+subrow, 7).value)) for subrow in range(1,9)]
                    self.VEHICLES_DATA_MP[(mode,product)].loc[int(t)] = values
                                        
            
            if ws.cell(row, 11).value is not None and type(ws.cell(row, 11).value) == str and ws.cell(row, 11).value.split(" ")[0] in self.F_FUEL:
                fuel = ws.cell(row, 11).value.split(" ")[0]
                t_start = int(ws.cell(row, 12).value)
                t_end = int(ws.cell(row, 13).value)
                for t in range(t_start, t_end+1):
                    values = [sigmoid(t,
                                      float(ws.cell(row+subrow, 12).value),
                                      float(ws.cell(row+subrow, 13).value),
                                      t_start, t_end,
                                      float(ws.cell(row+subrow, 15).value),
                                      float(ws.cell(row+subrow, 16).value),
                                      float(ws.cell(row+subrow, 17).value))
                              for subrow in range(1,10)]
                    
                    if t in self.VEHICLES_DATA_MF[(mode,fuel)].index:
                        for subrow in range(1, 10):
                            self.VEHICLES_DATA_MF[(mode,fuel)].loc[t, columns_MF[:10][subrow]] = values[subrow-1]
                    else:
                        self.VEHICLES_DATA_MF[(mode,fuel)].loc[int(t)] = values + 6*[0]

                    
            
            col = 23
            while ws.cell(row, col).value is not None and type(ws.cell(row, col).value) == str and ws.cell(row, col).value.split(" ")[-1][0].upper() + ws.cell(row, col).value.split(" ")[-1][1:] in self.F_FUEL:
                mode_bis = ws.cell(row, col).value.split(" ")[0].title()
                if mode_bis == "Truck":
                    mode_bis = "Road"
                fuel = ws.cell(row, col).value.split(" ")[-1][0].upper() + ws.cell(row, col).value.split(" ")[-1][1:]
                t_start = int(ws.cell(row, col+1).value)
                t_end = int(ws.cell(row, col+2).value)
                cur_mid = 12 #TODO:change hardcoded
                cur_a = 0.9
                cur_k = 0.4
                for t in range(t_start, t_end+1):
                    values = [sigmoid(t,
                                      float(ws.cell(row+subrow, col+1).value),
                                      float(ws.cell(row+subrow, col+2).value),
                                      t_start, t_end, cur_mid, cur_a, cur_k)
                              for subrow in range(1,7)]
                    
                    if t in self.VEHICLES_DATA_MF[(mode_bis,fuel)].index:
                        for subrow in range(1,7):
                            self.VEHICLES_DATA_MF[(mode_bis,fuel)].loc[t, columns_MF[10:][subrow]] = values[subrow-1]
                    else:
                        self.VEHICLES_DATA_MF[(mode_bis,fuel)].loc[int(t)] = 9*[0] + values
                
                col += 5
        
        # WARNING: fuel cost values inconsistent between table on right and left side of sheet input parameters 
        # Values here: left table. Values after next section: right values
        if True:
            for row in range(1, ws.max_row+1):
                if ws.cell(row, 9).value is None and row > 5:
                    break
                elif ws.cell(row, 9).value is not None:
                    for key in self.VEHICLES_DATA_MF.keys(): # key = ("Road", "Battery") for example
                        if ws.cell(row, 9).value in key or ws.cell(row, 9).value == "Electricity":
                            for col in range(10, 100):
                                if ws.cell(row, col).value is not None:
                                    if ws.cell(row, 9).value == "Electricity" and key[0] != "Sea":
                                        self.VEHICLES_DATA_MF[(key[0], "Battery")].loc[ws.cell(3, col).value, "Fuel Cost"] = ws.cell(row, col).value
                                        if key[0] == "Rail":
                                            self.VEHICLES_DATA_MF[(key[0], "Catenary")].loc[ws.cell(3, col).value, "Fuel Cost"] = ws.cell(row, col).value
                                        # print(key, ws.cell(3, col).value, self.VEHICLES_DATA_MF[key].loc[ws.cell(3, col).value, "Fuel Cost"])
                                    else:
                                        self.VEHICLES_DATA_MF[key].loc[ws.cell(3, col).value, "Fuel Cost"] = ws.cell(row, col).value
                                else:
                                    break
    
        for (m,f) in taxes.keys():
            self.VEHICLES_DATA_MF[(m,f)]["Fuel Cost"] = self.VEHICLES_DATA_MF[(m,f)]["Fuel Cost"]  + taxes[(m,f)]
            
        
    @timeit
    def construct_emission_transfer(self, transfer_data, CO2_fee_data, co2_fee):
        
        #------------------------
        "Parameters part 1"
        #-----------------------      

        #If this emission cap is too stringent, the problem becomes infeasible.
        self.EMISSION_CAP_RELATIVE = {2023: 100, 
                                      2026: 72.5, 
                                      2028: 60, 
                                      2030: 45, 
                                      2034: 40, 
                                      2040: 25, 
                                      2050: 10}  #note that the targets also are used for plotting purposes.
        self.EMISSION_CAP_ABSOLUTE_BASE_YEAR = None
        
        self.TRANSFER_COST_PER_MODE_PAIR = {}   # transfer cost per origin mode, dest mode, product class
        self.VEHICLE_TRANSFER_TIME_IN_HARBOR = {}   # average time a ship/train (and thus the cargo) spends in a harbour/terminal for loading/unloading purposes
        self.CARGO_WAITING_TIME_TERMINAL = {}   # the average time that cargo waits at the terminal when in transit
        self.TOTAL_TRANSFER_TIME = {}

        # extract transfer costs from table
        for index, row in transfer_data.iterrows():
            index = row["Index"]
            if str(index).isnumeric():
                orig = row["From"]
                dest = row["To"]
                prod_class = row["Product class"]
                transf_cost = row["Total cost (NOK/Tonne)"]
                transf_time = row["Transfer loading/unloading total [h]"]
                waiting_time = row["Waiting time cargo in port [h]"]
                self.TRANSFER_COST_PER_MODE_PAIR[(orig, dest, prod_class)] = round(transf_cost/self.scaling_factor_monetary*self.scaling_factor_weight,self.precision_digits)    # 10E6NOK/10E6TONNES
                self.TRANSFER_COST_PER_MODE_PAIR[(dest, orig, prod_class)] = round(transf_cost/self.scaling_factor_monetary*self.scaling_factor_weight,self.precision_digits)    # 10E6NOK/10E6TONNES
                #No immediate need for scaling, as already in NOK/Tonnes
                self.VEHICLE_TRANSFER_TIME_IN_HARBOR[(orig, dest, prod_class)] = transf_time
                self.VEHICLE_TRANSFER_TIME_IN_HARBOR[(dest, orig, prod_class)] = transf_time
                self.CARGO_WAITING_TIME_TERMINAL[(orig, dest, prod_class)] = waiting_time
                self.CARGO_WAITING_TIME_TERMINAL[(dest, orig, prod_class)] = waiting_time
                self.TOTAL_TRANSFER_TIME[(orig, dest, prod_class)] = transf_time+waiting_time
                self.TOTAL_TRANSFER_TIME[(dest, orig, prod_class)] = transf_time+waiting_time
        

        if True: 
            # read CO2 fee
            
            self.CO2_fee = {}     
            # for y in self.T_TIME_PERIODS:
            #     CO2_fee = sigmoid(y, CO2_fee_data['B4'].value, 
            #                             CO2_fee_data['C4'].value, 
            #                             CO2_fee_data['B3'].value, 
            #                             CO2_fee_data['C3'].value, 
            #                             CO2_fee_data['E4'].value, 
            #                             CO2_fee_data['F4'].value, 
            #                             CO2_fee_data['G4'].value)  # #UNIT: EURO/TonneCO2 =    EURO/(1000*1000 gCO2)
            #     CO2_fee_adj = CO2_fee*EXCHANGE_RATE_EURO_TO_NOK/(1000*1000)   #NOK/gCO2
            #     self.CO2_fee[y] = CO2_fee_adj/self.scaling_factor_monetary*self.scaling_factor_emissions

            row_years = 52
            if co2_fee == "base":
                row_tax = 53 
            elif co2_fee == "low":
                row_tax = 54 
            elif co2_fee == "high":
                row_tax = 55
            elif co2_fee == "intermediate":
                row_tax = 56 
            column_start = "J"
            index_start = column_index_from_string(column_start)
            column_end = "AK"
            index_end = column_index_from_string(column_end)
            for col in range(index_start, index_end + 1):
                t = CO2_fee_data[get_column_letter(col) + str(row_years)].value
                fee = CO2_fee_data[get_column_letter(col) + str(row_tax)].value #EURO/TonneCO2
                fee_adj = fee*EXCHANGE_RATE_EURO_TO_NOK/(1000*1000)   #NOK/gCO2
                self.CO2_fee[t] = fee_adj/self.scaling_factor_monetary*self.scaling_factor_emissions

        COST_BIG_M = 10**8
        #base level transport costs (in average scenario)
        self.C_TRANSP_COST_BASE = {(i,j,m,r, f, p, t): COST_BIG_M for (i,j,m,r) in self.A_ARCS for f in self.FM_FUEL[m] 
                              for p in self.P_PRODUCTS for t in self.T_TIME_PERIODS}   #UNIT: NOK/T
        #scenario-dependent transport cost (computed using base cost)
        self.C_TRANSP_COST_NORMALIZED = {(m,f, p, t): COST_BIG_M for m in self.M_MODES for f in self.FM_FUEL[m] 
                              for p in self.P_PRODUCTS for t in self.T_TIME_PERIODS}   #UNIT: NOK/Tkm
        self.E_EMISSIONS_NORMALIZED = {(m,f,p,t): COST_BIG_M for m in self.M_MODES for f in self.FM_FUEL[m] 
                            for p in self.P_PRODUCTS for t in self.T_TIME_PERIODS}      #UNIT:  gCO2/T
        self.C_TRANSP_COST = {(i,j,m,r, f, p, t,s): COST_BIG_M for (i,j,m,r) in self.A_ARCS for f in self.FM_FUEL[m] 
                              for p in self.P_PRODUCTS for t in self.T_TIME_PERIODS for s in self.S_SCENARIOS}   #UNIT: NOK/T
        self.E_EMISSIONS = {(i,j,m,r,f,p,t): COST_BIG_M for (i,j,m,r) in self.A_ARCS for f in self.FM_FUEL[m] 
                            for p in self.P_PRODUCTS for t in self.T_TIME_PERIODS}      #UNIT:  gCO2/T
        self.C_CO2 = {(i,j,m,r,f,p,t): COST_BIG_M for (i,j,m,r) in self.A_ARCS for f in self.FM_FUEL[m] 
                      for p in self.P_PRODUCTS for t in self.T_TIME_PERIODS}   #UNIT: nok/T
        # new: time value
        self.TIME_VALUE_PER_TH = {p: COST_BIG_M for p in self.P_PRODUCTS}   #UNIT: NOK/Th (NOK per tonne-hour)
        self.TIME_IN_TERMINAL = {m: 0 for m in self.M_MODES}   #UNIT: NOK/Th (NOK per tonne-hour)
        self.SPEED = {m: 0.0 for m in self.M_MODES}     # UNIT: KM/H
        self.C_TIME_VALUE = {(i,j,m,r,p): COST_BIG_M for (i,j,m,r) in self.A_ARCS for p in self.P_PRODUCTS}   #UNIT: NOK/T


        # read emissions
        for m in self.M_MODES:
            for f in self.FM_FUEL[m]:
                for pc in self.PC_PRODUCT_CLASSES:
                    if m != "Sea" or pc != "Liquid bulk":
                        for p in self.PC_TO_P[pc]:
                            for y in self.T_TIME_PERIODS:
                                cur_val = 1000 * self.VEHICLES_DATA_MF[(m,f)].loc[y, "CO2 emission per km"] \
                                    / (self.VEHICLES_DATA_MP[(m,pc)].loc[y, "Tonnage"]*self.VEHICLES_DATA_MP[(m,pc)].loc[y, "Market failure (av. utilization)"])
                                self.E_EMISSIONS_NORMALIZED[(m,f,p,y)] = round(cur_val * self.scaling_factor_weight/self.scaling_factor_emissions,self.precision_digits)
                    else:
                        for y in self.T_TIME_PERIODS:
                            self.E_EMISSIONS_NORMALIZED[(m,f,pc,y)] = round(0 * self.scaling_factor_weight/self.scaling_factor_emissions,self.precision_digits)
        
        
        
        # process emissions per arc
        for (i,j,m,r) in self.A_ARCS:
            a = (i, j, m, r)
            for f in self.FM_FUEL[m]:
                for p in self.P_PRODUCTS:
                    for y in self.T_TIME_PERIODS:
                        self.E_EMISSIONS[(i, j, m, r, f, p, y)] = round(self.AVG_DISTANCE[a] * self.E_EMISSIONS_NORMALIZED[(m,f,p,y)], self.precision_digits)
                        #CO2 costs per tonne:
                        self.C_CO2[(i, j, m, r, f, p, y)] =  round(self.E_EMISSIONS[(i, j, m, r, f, p, y)] * self.CO2_fee[y], self.precision_digits)
        

    # @timeit
    def get_cost_MFPA(self, f, p, a, y, TIMES_data=None):
        (i, j, m, r) = a
        
        def equation(wacc, lifetime, capex):
            return ((1+wacc)**int(lifetime))*wacc*capex/(((1+wacc)**int(lifetime))-1)
        
        cost_euro = 0 #euro/tkm
        wacc = self.VEHICLES_DATA_MF[(m,f)].loc[y, "WACC"]
        lt_veh = self.VEHICLES_DATA_MF[(m,f)].loc[y, "Lifetime vehicle"]
        lt_cus = self.VEHICLES_DATA_MP[(m,p)].loc[y, "Lifetime costumization"]
        tonnage = self.VEHICLES_DATA_MP[(m,p)].loc[y, "Tonnage"]
        mileage = self.VEHICLES_DATA_MP[(m,p)].loc[y, "Annual Mileage"]
        
        fuel_cost_i = fuel_cost_j = self.VEHICLES_DATA_MF[(m,f)].loc[y, "Fuel Cost"]
        # print(fuel_cost_i, self.VEHICLES_DATA_MF[(m,f)].loc[y, "Fuel Cost"])
        p_i = p_j = 0.5
        
        if TIMES_data != None:
            reg_i, p_i = self.sreg_to_treg[(m, i, j)][0]
            reg_j, p_j = self.sreg_to_treg[(m, i, j)][1]
            if p_i + p_j != 1:
                raise Exception(f"fuel cost not totally accounted for: {p_i}, {p_j}, {reg_i}, {reg_j}")
            
            for n, reg_x in enumerate([reg_i, reg_j]):
                if reg_x not in ["EUR", 0, "0", "SWE1", "SWE2", "SHE"]:
                    if (f, m, reg_x, y) in TIMES_data.prices_TIMES_fmrp.keys() and not np.isnan(TIMES_data.prices_TIMES_fmrp[(f, m, reg_x, y)]):
                        coeff = 1
                        if n == 0:
                            fuel_cost_i = TIMES_data.prices_TIMES_fmrp[(f, m, reg_i, y)] * coeff
                            # print(fuel_cost_i/self.VEHICLES_DATA_MF[(m,f)].loc[y, "Fuel Cost"], self.VEHICLES_DATA_MF[(m,f)].loc[y, "Fuel Cost"], fuel_cost_i)
                        elif n == 1:
                            fuel_cost_j = TIMES_data.prices_TIMES_fmrp[(f, m, reg_j, y)] * coeff
                            # print(fuel_cost_j/self.VEHICLES_DATA_MF[(m,f)].loc[y, "Fuel Cost"], self.VEHICLES_DATA_MF[(m,f)].loc[y, "Fuel Cost"], fuel_cost_j)
                    # else:
                    #     if y != 2023:
                    #         print(f, m, reg_x, y, "NOT IN TIMES")
                    #     if n == 0:
                    #         fuel_cost_i = 1000000
                    #     elif n == 1:
                    #         fuel_cost_j = 1000000
            self.df_cost.loc[len(self.df_cost.index)] = [reg_i, m, f, y, fuel_cost_i] 
            self.df_cost.loc[len(self.df_cost.index)] = [reg_j, m, f, y, fuel_cost_j]                         
        cost_euro += equation(wacc, lt_veh, self.VEHICLES_DATA_MF[(m,f)].loc[y, "Capex vehicle"])
        cost_euro += equation(wacc, lt_veh, self.VEHICLES_DATA_MF[(m,f)].loc[y, "Residual value (vehicle)"])
        cost_euro += equation(wacc, lt_cus, self.VEHICLES_DATA_MP[(m,p)].loc[y, "Capex costumization"])
        cost_euro += equation(wacc, lt_cus, self.VEHICLES_DATA_MP[(m,p)].loc[y, "Residual value (costumization)"])
        cost_euro += self.VEHICLES_DATA_MP[(m,p)].loc[y, "Opex fix (admin, crew, insurance)"]
        cost_euro += self.VEHICLES_DATA_MF[(m,f)].loc[y, "Opex maintenance & repair"]
        cost_euro += self.VEHICLES_DATA_MP[(m,p)].loc[y, "Opex var (mode-fees)"]*mileage
        cost_euro += self.VEHICLES_DATA_MF[(m,f)].loc[y, "Emission cost"]*mileage
        cost_euro += self.VEHICLES_DATA_MF[(m,f)].loc[y, "Fuel Economy"]\
                        *(fuel_cost_i * p_i + fuel_cost_j * p_j) * mileage
        
        if False:
            false_cost_euro = cost_euro-2*equation(wacc, lt_cus, self.VEHICLES_DATA_MP[(m,p)].loc[y, "Residual value (costumization)"])
        else:
            false_cost_euro = cost_euro
        
        cost_euro_per_t_km = cost_euro/(mileage*tonnage)
        cost_euro_per_t_km += false_cost_euro*(1/self.VEHICLES_DATA_MP[(m,p)].loc[y, "Market failure (av. utilization)"] - 1)/(mileage*tonnage)
        cost_euro_per_t_km += false_cost_euro*(1/(1-self.VEHICLES_DATA_MF[(m,f)].loc[y, "Payload loss (fuel)"]) - 1)/(mileage*tonnage)

        return cost_euro_per_t_km #euro/tkm
    
    @timeit
    def construct_costs(self, param_input, conv_file, TIMES_data=None):
        # read transport costs
        
        self.df_cost = pd.DataFrame(columns=["region", "mode", 'fuel', "year","price"])
        
        
        self.sreg_to_treg = dict()
        for (i,j,m,r) in self.A_ARCS:
            a = (i, j, m, r)
            if TIMES_data is not None:
                n_i = self.centroid_to_nr[i]
                n_j = self.centroid_to_nr[j]
                for index, row in conv_file.iterrows():
                    if row["Mode"] == m:
                        if row["From"] == n_i and row["To"] == n_j:
                            self.sreg_to_treg[(m, i, j)] = [[row["z1"], row["p1"]],
                                                            [row["z2"], row["p2"]]]
                            break
                        elif row["From"] == n_j and row["To"] == n_i:
                            self.sreg_to_treg[(m, i, j)] = [[row["z2"], row["p2"]],
                                                            [row["z1"], row["p1"]]]
                            break
            
            for f in self.FM_FUEL[m]:
                fg = self.F_TO_FG[f]
                                
                for p in self.P_PRODUCTS:
                    for y in self.T_TIME_PERIODS:   
                        cost_euro = self.get_cost_MFPA(f,self.P_TO_PC[p], a, y, TIMES_data)
                        # if f == "Hydrogen" and m == "Road":
                        #     print(p, y, a, cost_euro)
                        # if f == "Battery" and m == "Road":
                        #     print(p, y, a, cost_euro)
                        cost = cost_euro*EXCHANGE_RATE_EURO_TO_NOK
                        self.C_TRANSP_COST_NORMALIZED[(m,f,p,y)] = round(cost/self.scaling_factor_monetary*self.scaling_factor_weight,self.precision_digits)
                        #compute base cost
                        self.C_TRANSP_COST_BASE[(i, j, m, r, f, p, y)] = round((self.AVG_DISTANCE[a] * self.C_TRANSP_COST_NORMALIZED[(m,f,p,y)]), self.precision_digits) 
                        #^: MINIMUM 6.7, , median = 114.8, 90%quantile = 2562.9,  max 9.6*10^7!!!
                        # self.E_EMISSIONS[(i, j, m, r, f, p, y)] = round(self.AVG_DISTANCE[a] * self.E_EMISSIONS_NORMALIZED[(m,f,p,y)], self.precision_digits)
                        # #CO2 costs per tonne:
                        # self.C_CO2[(i, j, m, r, f, p, y)] =  round(self.E_EMISSIONS[(i, j, m, r, f, p, y)] * self.CO2_fee[y], self.precision_digits)
                        
                        for s in self.S_SCENARIOS: # define transport cost for each scenario
                            if y in self.T_TIME_FIRST_STAGE_BASE: #only update second-stage costs!
                                self.C_TRANSP_COST[(i, j, m, r, f, p, y, s)] = round(self.C_TRANSP_COST_BASE[(i, j, m, r, f, p, y)] * 1,self.precision_digits)
                            elif y in self.T_TIME_SECOND_STAGE_BASE:
                            #transport cost = base transport cost * cost factor for fuel group associated with (m,f) for current active scenario:
                                scen_nr = self.scenario_information.scen_name_to_nr[s]
                                fg_scen = self.scenario_information.fg_fuel_cost_path_name[scen_nr][fg] #  "B", "P" or "O"
                                pc = self.P_TO_PC[p]
                                self.C_TRANSP_COST[(i, j, m, r, f, p, y, s)] = round(self.C_TRANSP_COST_BASE[(i, j, m, r, f, p, y)] * 
                                                                                    self.scenario_information.cost_factor[(fg_scen,y,m,pc,f)],
                                                                                    #self.scenario_information.mode_fuel_cost_factor[self.scenario_information.scen_name_to_nr[s]][(m,f)],
                                                                                    self.precision_digits)
                        
                        
    @timeit
    def construct_time_value(self, time_value_data, speed_data):
        # read time values
        for index, row in time_value_data.iterrows():
            self.TIME_VALUE_PER_TH[row["Product group"]] = round(row["Time value (EUR/th)"]*EXCHANGE_RATE_EURO_TO_NOK/
                                                                 self.scaling_factor_monetary*self.scaling_factor_weight,self.precision_digits)
        
        # read mode speeds
        for index, row in speed_data.iterrows():
            self.SPEED[row["Mode"]] = row["Speed (km/h)"]

        # translate to time values per arc
        for (i,j,m,r) in self.A_ARCS:
            for p in self.P_PRODUCTS:
                travel_time = (self.DISTANCE[(i,j,m,r)] / self.SPEED[m]) 
                #terminal_time = self.TIME_IN_TERMINAL[m]   
                self.C_TIME_VALUE[(i,j,m,r,p)] = round((travel_time)*self.TIME_VALUE_PER_TH[p],self.precision_digits)  # dist / speed * cost per hour
                
    @timeit                                         
    def construct_investments(self, node_cap_data, edge_cap_data):
            
        #################
        #  INVESTMENTS  #
        #################
        
        # Edge upgrades
        self.E_EDGES_INV = []           # list of edges that can be invested in, i.e., that are capacitated
        self.E_EDGES_UPG = []           # list of upgradeable edges
        for index, row in edge_cap_data.iterrows():
            # define edge
            (i,j,m,r) = (row["from_centroid"], row["to_centroid"], row["Mode"], row["Route"])    
            edge = (i,j,m,r)
            if (i,j,m,r) not in self.E_EDGES:
                edge = (j,i,m,r)    # flip edge if necessary
            # add to E_EDGES_INV if possible to invest
            if row["Capacity (tonnes)"] != -1:
                if row["Capacity increase (tonnes)"] > 0:
                    self.E_EDGES_INV.append(edge)
            # add to E_EDGES_UPG if possible to upgrade              
            if row["Upgradeable"] == 1:
                self.E_EDGES_UPG.append(edge) 
        
        self.U_UPGRADE = []     # list of type of upgrades
        for e in self.E_EDGES_UPG:
            self.U_UPGRADE.append((e,'Catenary'))            # HARDCODED


        # Initialize capacities, investments, and costs

        # edges
        self.Q_EDGE_BASE = {}           # dict of initial edge capacities       # TONNES      
        self.Q_EDGE_INV = {}            # dict of possible edge investments     # TONNES 
        self.C_EDGE_INV = {}            # dict of edge investment costs         # NOK 
        self.C_EDGE_UPG = {}            # dict of edge upgrade costs            # NOK 
        self.LEAD_TIME_EDGE_INV = {}  # dict of edge investment lead times    # YEARS
        self.LEAD_TIME_EDGE_UPG = {}  # dict of edge upgrade lead times       # YEARS

        # nodes
        self.Q_NODE_BASE = {}           # dict of initial node capacities       # TONNES      
        self.Q_NODE_INV = {}            # dict of possible node investments     # TONNES 
        self.C_NODE_INV = {}            # dict of node investment costs         # NOK 
        self.LEAD_TIME_NODE_INV = {}  # dict of node investment lead times    # YEARS


        # fill edge data
        for index, row in edge_cap_data.iterrows():
            # define edge
            (i,j,m,r) = (row["from_centroid"], row["to_centroid"], row["Mode"], row["Route"])       
            edge = (i,j,m,r)
            if (i,j,m,r) not in self.E_EDGES:
                edge = (j,i,m,r)    # flip edge if necessary
            
            # initial capacities
            self.Q_EDGE_BASE[edge] = round(row["Capacity (tonnes)"]/self.scaling_factor_weight, self.precision_digits)

            # investments
            if row["Capacity (tonnes)"] != -1:
                self.Q_EDGE_INV[edge] = round(row["Capacity increase (tonnes)"]/self.scaling_factor_weight, self.precision_digits)
                self.C_EDGE_INV[edge] = round(row["Investment cost (NOK)"]/self.scaling_factor_monetary, self.precision_digits)
                self.LEAD_TIME_EDGE_INV[edge] = row["Lead time (years)"]

            # upgrades
            if row["Upgradeable"] == 1:
                self.C_EDGE_UPG[(edge, 'Catenary')] = round(row["Upgrade cost (NOK)"]/self.scaling_factor_monetary,self.precision_digits)                 # HARDCODED
                self.LEAD_TIME_EDGE_UPG[(edge, 'Catenary')] = row["Upgrade lead time (years)"]  # HARDCODED
        
        self.NM_CAP = []
        self.NM_CAP_INCR = []
        # fill node data
        for index, row in node_cap_data.iterrows():
            # define node
            i = row["centroid_name"]
            m = row["Mode"]
            
            
            # investments
            if row["Capacity (tonnes)"] != -1:
                
                # initial capacities
                self.Q_NODE_BASE[(i,m)] = round(row["Capacity (tonnes)"]/self.scaling_factor_weight, self.precision_digits)
                self.Q_NODE_INV[(i,m)] = round(row["Capacity increase (tonnes)"]/self.scaling_factor_weight, self.precision_digits)
                self.C_NODE_INV[(i,m)] = round(row["Investment cost (NOK)"]/self.scaling_factor_monetary, self.precision_digits)
                self.LEAD_TIME_NODE_INV[(i,m)] = row["Lead time (years)"]
                self.NM_CAP.append((i,m))
                if row["Capacity increase (tonnes)"] > 0.01:
                    self.NM_CAP_INCR.append((i,m))
        

        # Big M        
        self.BIG_M_UPG = {e: [] for e in self.E_EDGES_UPG}        # TONNES 
        for e in self.E_EDGES_UPG:
            if e in self.Q_EDGE_INV:
                self.BIG_M_UPG[e] =  1.5*(self.Q_EDGE_BASE[e] + self.Q_EDGE_INV[e])
            else:
                self.BIG_M_UPG[e] =  1.5*self.Q_EDGE_BASE[e]


        # Discount rate
        self.risk_free_interest_rate = RISK_FREE_RATE # 2%
        self.D_DISCOUNT_RATE = round(1 / (1 + self.risk_free_interest_rate),self.precision_digits)
        
    @timeit
    def construct_charging_edges(self, charging_data):
        # --------------------------
        # CHARGING EDGES CONSTRAINT
        # --------------------------
        
        self.CHARGING_TECH = []
        T_TIME_PERIODS_CHARGING = set()
        for index,row in charging_data.iterrows():
            #print((row["Mode"],row["Fuel"]))
            self.CHARGING_TECH.append((row["Mode"],row["Fuel"]))
            T_TIME_PERIODS_CHARGING.add(row["Year"])
        self.CHARGING_TECH = list(set(self.CHARGING_TECH))

        # all arcs (one per arc pair ij/ji) with mode Road and fuels Battery or Hydrogen
        self.EF_CHARGING = []
        for (i,j,m,r) in self.E_EDGES:
            e =(i,j,m,r)
            if i not in self.N_ABROAD or j not in self.N_ABROAD: #and or 'or'
                for (m, f) in self.CHARGING_TECH:
                    if e[2] == m:
                        self.EF_CHARGING.append((e,f))
        
        # base capacity on a pair of arcs (ij/ji - mfr), fix to 0 since no charging infrastructure exists now
        self.Q_CHARGE_BASE = {(e,f): 0 for (e,f) in self.EF_CHARGING }
        self.C_CHARGE = {(e,f,t): 100000 for (e,f) in self.EF_CHARGING for t in T_TIME_PERIODS_CHARGING}  # for p in self.P_PRODUCTS}  
        #self.LEAD_TIME_CHARGING = {(e,f,t): 50 for (e,f) in self.EF_CHARGING for t in self.charging_years}
        avg_payload = AVG_TRUCK_PAYLOAD  # HARDCODE random average in tonnes, should be product based? or fuel based??
        
        for ((i, j, m, r),f) in self.EF_CHARGING:
            e = (i, j, m, r)
            for  t in T_TIME_PERIODS_CHARGING:
                data_index = charging_data.loc[(charging_data['Mode'] == m) & (charging_data['Fuel'] == f) & (charging_data['Year'] == t)]  #pick the right row index
                num_stations_on_arc = self.AVG_DISTANCE[e]/ data_index.iloc[0]["Max_station_dist_km"]
                self.C_CHARGE[(e,f,t)] = round((num_stations_on_arc* data_index.iloc[0]["Station_cost_NOK"]  #annualized
                                                    / (data_index.iloc[0]["Trucks_filled_daily"] * avg_payload * 365) #do scaling of necessary number of stations to get the cost per tonne per year.
                                                    /self.scaling_factor_monetary
                                                    *self.scaling_factor_weight),
                                            self.precision_digits)  # 0.7 or not??? MKR/TONNES, 
                #self.LEAD_TIME_CHARGING[(e,f,t)] = data_index.iloc[0]["Ledetid_year"]

    @timeit
    def construct_tech_readiness(self, tech_readiness_data, phase_out_data, init_transport_share, init_mode_share):
        ##################################
        #    Technological readiness
        ##################################

        #Technological readiness/maturity (with Bass diffusion model)
        
        self.tech_is_mature = {} # indicates whether technology is already mature
        self.tech_base_bass_model = {} # contains all bass diffusion models for the technologies (only for non-mature technologies)
        self.tech_active_bass_model = {} # active bass model (only for non-mature technologies)
        self.tech_scen_p_q_variation = {} # variation (as %) for parameters p and q in the bass diffusion model in each of the scenarios
        self.tech_scen_t_0_delay = {} # delay for parameter in the bass diffusion model in each of the scenarios
        for index, row in tech_readiness_data.iterrows():
            # store whether technology is already mature or not 
            if row["Mature?"] == "yes":
                self.tech_is_mature[(row['Mode'], row['Fuel'])] = True
            else:
                self.tech_is_mature[(row['Mode'], row['Fuel'])] = False
                # if not mature, add bass diffusion model
                self.tech_base_bass_model[(row['Mode'], row['Fuel'])] = BassDiffusion(float(row["p"]), float(row["q"]), float(row["m"]), int(row["t_0"]))
                # set base bass model as active bass model
                #self.tech_active_bass_model[(row['Mode'] ,row['Fuel'])] = BassDiffusion(float(row["p"]), float(row["q"]), float(row["m"]), int(row["t_0"]))
                # store variations
                self.tech_scen_p_q_variation[(row['Mode'], row['Fuel'])] = row["p_q_variation"]
                self.tech_scen_t_0_delay[(row['Mode'], row['Fuel'])] = row["t_0_delay"]
        
        self.R_TECH_READINESS_MATURITY = {} # contains the active maturity path (number between 0 and 100)
        # initialize R_TECH_READINESS_MATURITY at base path
        for s in self.S_SCENARIOS:
            for (m,f) in self.tech_is_mature:
                if self.tech_is_mature[(m,f)]:
                    for year in self.T_TIME_PERIODS:    
                        self.R_TECH_READINESS_MATURITY[(m, f, year,s)] = 100 # assumption: all mature technologies have 100% market potential
                else:
                    for year in self.T_TIME_PERIODS:
                        #we can remove this one!
                        self.R_TECH_READINESS_MATURITY[(m, f, year,s)] = round(self.tech_base_bass_model[(m,f)].A(year),self.precision_digits) # compute maturity level based on base Bass diffusion model 
        
        #Phase out technologies
        self.PHASE_OUT = {}
        for index, row in phase_out_data.iterrows():
            m=row["Mode"]
            f=row["Fuel"]
            t=row["Time period"]
            self.PHASE_OUT[(m,f,t)] = row["Restriction"]

        #Initializing transport work share in base year
        
        self.Q_SHARE_INIT_MAX = {}
        self.MFT_INIT_TRANSP_SHARE = []
        for index, row in init_transport_share.iterrows():
            (m,f,t) = (row['Mode'], row['Fuel'],self.T_TIME_PERIODS[0])
            self.Q_SHARE_INIT_MAX[(m,f,t)] = round(row['Max_transp_share'],self.precision_digits)
            self.MFT_INIT_TRANSP_SHARE.append((m,f,t))

        
        self.INIT_MODE_SPLIT = {m:None for m in self.M_MODES}
        for index, row in init_mode_share.iterrows():
            (mm,share) = (row['Mode'], row['Share'])
            self.INIT_MODE_SPLIT[mm] = round(share,self.precision_digits)

        #update R_TECH_READINESS_MATURITY based on scenario information
        # TURNED OFF FOR NOW
        if False:
            for s in self.S_SCENARIOS:
                active_scenario_nr = self.scenario_information.scen_name_to_nr[s]
                for m in self.M_MODES:
                    for f in self.FM_FUEL[m]:
                        if not self.tech_is_mature[(m,f)]: # only vary maturity information by scenario for non-mature technologies
                            cur_fg = self.scenario_information.mf_to_fg[(m,f)]
                            cur_path_name = self.scenario_information.fg_maturity_path_name[active_scenario_nr][cur_fg] # find name of current maturity path [base, fast, slow]
                            # extract info from current base Bass model
                            cur_base_bass_model = self.tech_base_bass_model[(m,f)] # current base Bass diffusion model
                            cur_base_p_q_variation = self.tech_scen_p_q_variation[(m,f)] # level of variation for this m,f 
                            cur_base_t_0_delay = self.tech_scen_t_0_delay[(m,f)] # time delay for t_0 for this m,f
                                        
                            # find current scenario's level of variation for q and p and delay for t_0 from base case
                            cur_scen_p_q_variation = 0.0 
                            cur_scen_t_0_delay = 0.0
                            if cur_path_name == "B":
                                cur_scen_p_q_variation = 0.0
                                cur_scen_t_0_delay = 0.0
                            if cur_path_name == "O":
                                cur_scen_p_q_variation = cur_base_p_q_variation # increase p and q by cur_base_p_q_variation (e.g., 50%)
                                cur_scen_t_0_delay = - cur_base_t_0_delay # negative delay (faster development)
                            elif cur_path_name == "P":
                                cur_scen_p_q_variation = - cur_base_p_q_variation # decrease p and q by cur_base_p_q_variation (e.g., 50%)
                                cur_scen_t_0_delay = cur_base_t_0_delay # positive delay (slower development)

                            # construct scenario bass model
                            cur_scen_bass_model = BassDiffusion(cur_base_bass_model.p * (1 + cur_scen_p_q_variation), # adjust p with cur_scen_variations
                                                                cur_base_bass_model.q * (1 + cur_scen_p_q_variation),     # adjust q with cur_scen_variations
                                                                cur_base_bass_model.m, 
                                                                cur_base_bass_model.t_0 + cur_scen_t_0_delay)
                            
                            # set as active bass model
                            self.tech_active_bass_model[(m,f,s)] = cur_scen_bass_model

                            # find start of second stage
                            for t in self.T_TIME_PERIODS:
                                if t not in self.T_TIME_FIRST_STAGE_BASE:
                                    start_of_second_stage = t
                                    break

                            # fill R_TECH_READINESS_MATURITY based on current scenario bass model
                            for t in self.T_TIME_PERIODS:
                                if t in self.T_TIME_FIRST_STAGE_BASE:
                                    # first stage: follow base bass model
                                    self.R_TECH_READINESS_MATURITY[(m,f,t,s)] = round(cur_base_bass_model.A(t),self.precision_digits)
                                else:
                                    # second stage: use scenario bass model, with starting point A(2030) from base bass model
                                    t_init = start_of_second_stage #initialize diffusion at start of second stage
                                    A_init = cur_base_bass_model.A(t_init) # diffusion value at start of second stage 
                                    self.R_TECH_READINESS_MATURITY[(m,f,t,s)] = round(cur_scen_bass_model.A_from_starting_point(t,A_init,t_init),self.precision_digits)
        else:
            # we need to set the active bass model to the base bass model
            for s in self.S_SCENARIOS:
                for m in self.M_MODES:
                    for f in self.FM_FUEL[m]:
                        if self.tech_is_mature[(m,f)] == False:
                            self.tech_active_bass_model[(m,f,s)] = self.tech_base_bass_model[(m,f)]


    @timeit
    def construct_path_generation(self, path):
        #----------------------------------------
        #      PATH GENERATION
        #-----------------------------------------

        self.K_PATHS = []

        #from Data.settings import *
        
        if not os.path.exists(path):         
            path_generation(
                    products=self.P_PRODUCTS, 
                    p_to_pc=self.P_TO_PC,
                    modes=self.M_MODES, 
                    nodes=self.N_NODES, 
                    edges=self.E_EDGES, 
                    years=self.T_TIME_PERIODS, 
                    distances=self.AVG_DISTANCE,
                    transp_costs = self.C_TRANSP_COST_NORMALIZED,     
                    emissions=self.E_EMISSIONS_NORMALIZED,              
                    emission_fee=self.CO2_fee,           
                    prod_transfer_costs=self.TRANSFER_COST_PER_MODE_PAIR,    
                    mode_to_fuels=self.FM_FUEL,           
                    mode_comb_level=NUM_MODE_PATHS         
                    )

        #import ast
        #all_generated_paths = pd.read_csv(r'Data/SPATIAL/'+filename, converters={'paths': ast.literal_eval})  #, converters={'paths': eval}        #This provides an encoding error
        
        with open(path, 'rb') as file:
            all_generated_paths = pickle.load(file)
        
        self.K_PATH_DICT = {i:None for i in range(len(all_generated_paths))}
        #for index, row in all_generated_paths.iterrows():
        #    elem = tuple(row['paths']) 
        #    self.K_PATHS.append(index)
        #    self.K_PATH_DICT[index]=elem
        for i in range(len(all_generated_paths)):
            elem = tuple(all_generated_paths[i]) 
            self.K_PATHS.append(i)
            self.K_PATH_DICT[i]=elem        

        self.OD_PATHS = {od: [] for od in self.OD_PAIRS_ALL}
        for od in self.OD_PAIRS_ALL:
            for k in self.K_PATHS:
                path = self.K_PATH_DICT[k]
                if od[0] == path[0][0] and od[-1] == path[-1][1]:
                    self.OD_PATHS[od].append(k)

        #multi-mode paths and unimodal paths
        self.MULTI_MODE_PATHS = []
        for kk in self.K_PATHS:
            k = self.K_PATH_DICT[kk]
            if len(k) > 1:
                for i in range(len(k)-1):
                    if k[i][2] != k[i+1][2]:
                        self.MULTI_MODE_PATHS.append(kk)
        self.UNI_MODAL_PATHS = list(set(self.K_PATHS)-set(self.MULTI_MODE_PATHS))

        self.UNI_MODAL_PATHS_PER_MODE = {m:[] for m in self.M_MODES}
        for kk in self.UNI_MODAL_PATHS:
            (i,j,m,r) = self.K_PATH_DICT[kk][0]
            self.UNI_MODAL_PATHS_PER_MODE[m].append(kk)

        self.PATHS_NO_UNIMODAL_ROAD = list(set(self.K_PATHS)-set(self.UNI_MODAL_PATHS_PER_MODE["Road"]))


        #Paths with transfer in node i to/from mode m
        #self.TRANSFER_PATHS = {(i,m) : [] for m in self.M_MODES_CAP for i in self.N_NODES_CAP_NORWAY[m]}
        self.TRANSFER_PATHS = {(i,m) : [] for (i,m) in self.NM_CAP}  #When transfering, you will have to use a mode with a capacitated terminal (rail or sea)

        for (i,m) in self.NM_CAP:
            for kk in self.MULTI_MODE_PATHS:
                k = self.K_PATH_DICT[kk]
                for j in range(len(k)-1):
                    if (k[j][1] == i) and (k[j][2] == m or k[j+1][2] == m) and (k[j][2] != k[j+1][2]):
                        self.TRANSFER_PATHS[(i,m)].append(kk)

        #Origin and destination paths  (this is used to calculate the capacity usage in terminals)
        self.ORIGIN_PATHS = {(i,m): [] for (i,m) in self.NM_CAP}
        self.DESTINATION_PATHS = {(i,m): [] for (i,m) in self.NM_CAP}
        
        for (i,m) in self.NM_CAP:
            for kk in self.K_PATHS:
                k = self.K_PATH_DICT[kk]
                if (k[0][0] == i) and (k[0][2] == m):
                    self.ORIGIN_PATHS[(i,m)].append(kk)
                if (k[-1][1] == i) and (k[-1][2] == m):
                    self.DESTINATION_PATHS[(i,m)].append(kk)
        
        
        self.KA_PATHS = {a:[] for a in self.A_ARCS}
        for kk in self.K_PATHS:
            k = self.K_PATH_DICT[kk]
            for (i,j,m,r) in k:
                a = (i,j,m,r)
                self.KA_PATHS[a].append(kk)

        self.KA_PATHS_UNIMODAL = {a:[] for a in self.A_ARCS}
        for kk in self.UNI_MODAL_PATHS:
            k = self.K_PATH_DICT[kk]
            for (i,j,m,r) in k:
                a = (i,j,m,r)
                self.KA_PATHS_UNIMODAL[a].append(kk)
        
        self.FM_MULTI = {}
        for m1 in self.M_MODES:
            self.FM_MULTI[(m1,m1)] = []
            for f1 in self.FM_FUEL[m1]:
                self.FM_MULTI[(m1,m1)].append((f1,f1))
            for m2 in self.M_MODES:
                if m1 != m2:
                    self.FM_MULTI[(m1,m2)] = []
                    for f1 in self.FM_FUEL[m1]:
                        for f2 in self.FM_FUEL[m2]:
                            self.FM_MULTI[(m1,m2)].append((f1, f2))
        # For each modes combination, we have a list of possible fuel combination
        # Ex: self.FM_MULTI[("Road", "Rail")] = [("Diesel", "Catenary"), ("Diesel", "Battery"), ..., ("Hydrogen", "Catenary"),("Hydrogen", "Battery"), ...]
        self.FM_MULTI_K = {}
        for k in self.K_PATHS:
            result = [i[2] for i in self.K_PATH_DICT[k]]
            modes = [] 
            [modes.append(x) for x in result if x not in modes] 
            modes = tuple(modes)
            if len(modes) == 1:
                modes = modes + modes
            self.FM_MULTI_K[k] = self.FM_MULTI[modes]

    @timeit
    def construct_param2(self):
        #----------------------------------------
        #      Parameters part 2
        #-----------------------------------------

        # self.C_TRANSFER = {(k,p):0 for k in self.K_PATHS for p in self.P_PRODUCTS}   #UNIT: NOK/T     MANY ELEMENTS WILL BE ZERO!! (NO TRANSFERS)
        # for kk in self.MULTI_MODE_PATHS:
        #     k = self.K_PATH_DICT[kk]
        #     for p in self.P_PRODUCTS:
        #         cost = 0
        #         num_transfers = len(k)-1
        #         for n in range(num_transfers):
        #             mode_from = k[n][2]
        #             mode_to = k[n+1][2]
        #             if mode_from != mode_to: 
        #                 cost += self.TRANSFER_COST_PER_MODE_PAIR[mode_from, mode_to, self.P_TO_PC[p]]
        #         self.C_TRANSFER[(kk,p)] = round(cost,self.precision_digits)

        self.C_TRANSFER = {(k,p):0 for k in self.K_PATHS for p in self.P_PRODUCTS}   #UNIT: NOK/T     MANY ELEMENTS WILL BE ZERO!! (NO TRANSFERS)
        self.C_TRANSFER_TIME = {(k,p):0 for k in self.K_PATHS for p in self.P_PRODUCTS}
        for kk in self.PATHS_NO_UNIMODAL_ROAD:
            k = self.K_PATH_DICT[kk]
            for p in self.P_PRODUCTS:
                cost = 0
                time_cost = 0
                num_arcs = len(k)
                initial_mode = k[0][2]
                final_mode = k[num_arcs-1][2]
                if initial_mode in ["Rail", "Sea"]: #first mile with Road and hence a transfer cost
                    cost += self.TRANSFER_COST_PER_MODE_PAIR["Road", initial_mode, self.P_TO_PC[p]]
                    time_cost += self.TOTAL_TRANSFER_TIME["Road", initial_mode, self.P_TO_PC[p]]*self.TIME_VALUE_PER_TH[p]
                if final_mode in ["Rail", "Sea"]: #last mile with Road and hence a transfer cost 
                    cost += self.TRANSFER_COST_PER_MODE_PAIR[final_mode,"Road", self.P_TO_PC[p]]
                    time_cost += self.TOTAL_TRANSFER_TIME["Road", final_mode, self.P_TO_PC[p]]*self.TIME_VALUE_PER_TH[p]
                if num_arcs>1: #Calculate the transfer costs DURING the path (not first-/last-mile)
                    for n in range(num_arcs-1):
                        mode_from = k[n][2]
                        mode_to = k[n+1][2]
                        if mode_from != mode_to: 
                            cost += self.TRANSFER_COST_PER_MODE_PAIR[mode_from, mode_to, self.P_TO_PC[p]]
                            time_cost += self.TOTAL_TRANSFER_TIME[mode_from, mode_to, self.P_TO_PC[p]]*self.TIME_VALUE_PER_TH[p]
                self.C_TRANSFER[(kk,p)] = round(cost,self.precision_digits)
                self.C_TRANSFER_TIME[(kk,p)] = round(time_cost,self.precision_digits)
        
        #EMISSIONS
        #if EMISSION_CONSTRAINT:
        #    emission_cap_data = pd.read_excel(r'Data/transport_costs_emissions_raw.xlsx', sheet_name='emission_cap')

    
    @timeit
    def combined_sets(self, TIMES_data=None):

        if True:  #Q: is this still necessary? Or can we remove this
        
            #####################
            ## VEHICLE TYPES ####
            #####################
            
            #Vehicle types
            prod_to_vehicle_type = pd.read_excel(r'Data/transport_costs_emissions_raw.xlsx', sheet_name='prod_to_vehicle')
            self.VEHICLE_TYPE_MP = {}
            self.VEHICLE_TYPES_M = {m:[] for m in self.M_MODES}
            prod_veh_data = zip(prod_to_vehicle_type['Mode'], prod_to_vehicle_type['Product class'], prod_to_vehicle_type['Vehicle type'])
            for (m,pc,v) in prod_veh_data:
                for p in self.PC_TO_P[pc]:
                    if p in self.P_PRODUCTS:
                        self.VEHICLE_TYPE_MP[(m,p)] = v
                        self.VEHICLE_TYPES_M[m].append(v)
            for m in self.M_MODES:
                self.VEHICLE_TYPES_M[m] = list(set(self.VEHICLE_TYPES_M[m]))
            self.V_VEHICLE_TYPES = list(set(self.VEHICLE_TYPE_MP.values()))
            
            self.PV_PRODUCTS = {v:[] for v in self.V_VEHICLE_TYPES}
            for (m,p),v in self.VEHICLE_TYPE_MP.items():
                self.PV_PRODUCTS[v].append(p)


        self.SS_SCENARIOS_NONANT = []
        for s in self.S_SCENARIOS:
            for ss in self.S_SCENARIOS:
                if (s != ss) and ((ss,s) not in self.SS_SCENARIOS_NONANT):
                    self.SS_SCENARIOS_NONANT.append((s,ss))

        self.T_TIME_FIRST_STAGE = [t for t in self.T_TIME_FIRST_STAGE_BASE if t in self.T_TIME_PERIODS]  
        self.T_TIME_SECOND_STAGE = [t for t in self.T_TIME_SECOND_STAGE_BASE if t in self.T_TIME_PERIODS]  

        start = self.T_TIME_PERIODS[0]
        #if len(self.T_TIME_PERIODS) == len(self.T_TIME_PERIODS_ALL):
        end = self.T_TIME_PERIODS[len(self.T_TIME_PERIODS)-1] + 1   #we only need to model the development until the end
        #else:
        #    end = 
        self.T_YEARLY_TIME_PERIODS = [*range(start,end)] #all years from 2022 up to 2050
        self.T_YEARLY_TIME_PERIODS_ALL = [*range(start,self.T_TIME_PERIODS_ALL[len(self.T_TIME_PERIODS_ALL)-1] + 1)] #all years from 2022 up to 2050
        
        self.T_YEARLY_TIME_FIRST_STAGE = [ty for ty in self.T_YEARLY_TIME_PERIODS if ty < self.T_TIME_SECOND_STAGE_BASE[0] ]
        #self.T_YEARLY_TIME_FIRST_STAGE_NO_TODAY = [*range(self.T_TIME_PERIODS[0] + 1, 2030)] #first-stage years without the first period
        self.T_YEARLY_TIME_SECOND_STAGE = [ty for ty in self.T_YEARLY_TIME_PERIODS if ty >= self.T_TIME_SECOND_STAGE_BASE[0] ]
        
        self.Y_YEARS = {t:[] for t in self.T_TIME_PERIODS_ALL}
        t0 = self.T_TIME_PERIODS[0]
        num_periods = len(self.T_TIME_PERIODS_ALL)
        
        for i in range(num_periods):
            t = self.T_TIME_PERIODS_ALL[i]
            if i < num_periods-1:
                tp1 = self.T_TIME_PERIODS_ALL[i+1]
                self.Y_YEARS[t] = list(range(t-t0,tp1-t0))
            elif i == (num_periods - 1):  #this is the last time period. Lasts only a year?? 
                duration_previous = len(self.Y_YEARS[self.T_TIME_PERIODS_ALL[i-1]])
                self.Y_YEARS[t] = [self.T_TIME_PERIODS_ALL[i]-t0 + j for j in range(duration_previous)]

        self.T_MOST_RECENT_DECISION_PERIOD = {}
        for ty in self.T_YEARLY_TIME_PERIODS: #loop over all (yearly) years
            cur_most_recent_dec_period = self.T_TIME_PERIODS[0] #initialize at 2022
            for t in self.T_TIME_PERIODS: # loop over all decision periods
                if t <= ty:
                    cur_most_recent_dec_period = t 
            self.T_MOST_RECENT_DECISION_PERIOD[ty] = cur_most_recent_dec_period


        self.T_TIME_PERIODS_OPERATIONAL = self.T_TIME_PERIODS
        if self.single_time_period is not None:
            self.T_TIME_PERIODS_OPERATIONAL = [self.single_time_period]

        #
        #       WITHOUT SCENARIOS
        #
        
        #------------------------
        "Combined sets - time independent"
        #------------------------

        self.MF = [(m,f) for m in self.M_MODES for f in self.FM_FUEL[m]]

        "Combined sets - time dependent"

        self.TS = [(t,) for t in self.T_TIME_PERIODS]
        self.TS_CONSTR = [(t,) for t in self.T_TIME_PERIODS_OPERATIONAL]
        self.TS_NO_BASE_YEAR = [(t,) for t in self.T_TIME_PERIODS if t is not self.T_TIME_PERIODS[0]]
        self.TS_NO_BASE_YEAR_CONSTR = [(t,) for t in self.T_TIME_PERIODS_OPERATIONAL if t is not self.T_TIME_PERIODS[0]]


        self.APT = [(i,j,m,r) + (p,) + (t,) for (i,j,m,r) in self.A_ARCS for p in self.P_PRODUCTS for t in self.T_TIME_PERIODS] 
        self.AVT = [(i,j,m,r) + (v,) + (t,) for (i,j,m,r) in self.A_ARCS for v in self.VEHICLE_TYPES_M[m] for t in self.T_TIME_PERIODS] 
        self.APT_CONSTR = [(i,j,m,r) + (p,) + (t,) for (i,j,m,r) in self.A_ARCS for p in self.P_PRODUCTS for t in self.T_TIME_PERIODS_OPERATIONAL] 
        self.AVT_CONSTR = [(i,j,m,r) + (v,) + (t,) for (i,j,m,r) in self.A_ARCS for v in self.VEHICLE_TYPES_M[m] for t in self.T_TIME_PERIODS_OPERATIONAL] 
        
        
        self.AFPT = [(i,j,m,r) + (f,) + (p,) + (t,)  for (i,j,m,r) in self.A_ARCS for f in self.FM_FUEL[m] for p in self.P_PRODUCTS for t in
                         self.T_TIME_PERIODS ]
        self.AFVT = [(i,j,m,r) + (f,) + (v,) + (t,) for (i,j,m,r) in self.A_ARCS for f in self.FM_FUEL[m] for v in self.VEHICLE_TYPES_M[m] for t in
                         self.T_TIME_PERIODS ]  
        self.KPT = [(k, p, t) for k in self.K_PATHS for p in self.P_PRODUCTS for t in self.T_TIME_PERIODS]        
        self.KVT = [(k, v, t) for k in self.K_PATHS for v in self.V_VEHICLE_TYPES for t in self.T_TIME_PERIODS]
        self.K_uni_VT = [(k, v, t) for k in self.UNI_MODAL_PATHS for v in self.V_VEHICLE_TYPES for t in self.T_TIME_PERIODS]

        self.ET_INV = [l+(t,) for l in self.E_EDGES_INV for t in self.T_TIME_PERIODS                                   if (t <= self.T_TIME_PERIODS[-1] - self.LEAD_TIME_EDGE_INV[l]) and (t in self.T_TIME_FIRST_STAGE)]
        self.EAT_INV = [e+(a,)+(t,) for e in self.E_EDGES_INV for a in self.AE_ARCS[e] for t in self.T_TIME_PERIODS   if (t <= self.T_TIME_PERIODS[-1] - self.LEAD_TIME_EDGE_INV[e]) and (t in self.T_TIME_FIRST_STAGE)]
        self.EAT_INV_CONSTR = [e+(a,)+(t,) for e in self.E_EDGES_INV for a in self.AE_ARCS[e] for t in self.T_TIME_PERIODS_OPERATIONAL]        
        self.EFT_CHARGE = [(e,f,t) for (e,f) in self.EF_CHARGING for t in self.T_TIME_PERIODS ] # if t <= self.T_TIME_PERIODS[-1] - self.LEAD_TIME_CHARGING[(e,f)]]
        self.EFT_CHARGE_CONSTR = [(e,f,t) for (e,f) in self.EF_CHARGING for t in self.T_TIME_PERIODS_OPERATIONAL]
        self.NM_CAP_INCR_T = [(i,m,t) for (i,m) in self.NM_CAP_INCR for t in self.T_TIME_PERIODS if t <= self.T_TIME_PERIODS[-1] - self.LEAD_TIME_NODE_INV[i,m]]
        self.NM_CAP_T = [(i,m,t) for (i,m) in self.NM_CAP for t in self.T_TIME_PERIODS if t <= self.T_TIME_PERIODS[-1] - self.LEAD_TIME_NODE_INV[i,m]]
        self.NM_CAP_T_CONSTR = [(i,m,t) for (i,m) in self.NM_CAP for t in self.T_TIME_PERIODS_OPERATIONAL ]
        self.NMFVT = [(i,m,f,v,t) for m in self.M_MODES for f in self.FM_FUEL[m] for i in self.NM_NODES[m]
                                    for v in self.VEHICLE_TYPES_M[m] for t in self.T_TIME_PERIODS]
        self.NMFVT_CONSTR = [(i,m,f,v,t) for m in self.M_MODES for f in self.FM_FUEL[m] for i in self.NM_NODES[m]
                                    for v in self.VEHICLE_TYPES_M[m] for t in self.T_TIME_PERIODS_OPERATIONAL]
        self.ODPTS = [odp + (t,) for odp in self.ODP for t in self.T_TIME_PERIODS]
        self.ODPTS_CONSTR = [odp + (t,) for odp in self.ODP for t in self.T_TIME_PERIODS_OPERATIONAL]
        self.EPT = [l + (p,) + (t,) for l in self.E_EDGES for p in self.P_PRODUCTS for t in
                         self.T_TIME_PERIODS]
        self.MFT_MATURITY = [mf + (t,) for mf in self.NEW_MF_LIST for t in self.T_TIME_PERIODS]
        self.MFT_MATURITY_CONSTR = [mf + (t,) for mf in self.NEW_MF_LIST for t in self.T_TIME_PERIODS_OPERATIONAL]
        self.MFT = [(m,f,t) for m in self.M_MODES for f in self.FM_FUEL[m] for t in self.T_TIME_PERIODS]
        self.MFT_CONSTR = [(m,f,t) for m in self.M_MODES for f in self.FM_FUEL[m] for t in self.T_TIME_PERIODS_OPERATIONAL]
        
        self.MFTT = [(m,f,t,tau) for m in self.M_MODES for f in self.FM_FUEL[m] for t in self.T_TIME_PERIODS 
                            for tau in self.T_TIME_PERIODS if tau <= t]
        self.MFT_MIN0 = [(m,f,t) for m in self.M_MODES for f in self.FM_FUEL[m] 
                                    for t in self.T_TIME_PERIODS if t!=self.T_TIME_PERIODS[0]]
        self.MT_MIN0 = [(m,t) for m in self.M_MODES for t in self.T_TIME_PERIODS if t!=self.T_TIME_PERIODS[0]]

        self.MT = [(m,t) for m in self.M_MODES for t in self.T_TIME_PERIODS]

        self.MFT_NEW = [(m,f,t) for m in self.M_MODES for f in self.FM_FUEL[m] for t in self.T_TIME_PERIODS if not self.tech_is_mature[(m,f)]]
        self.MFT_NEW_YEARLY = [(m,f,t) for m in self.M_MODES for f in self.FM_FUEL[m] for t in self.T_YEARLY_TIME_PERIODS if not self.tech_is_mature[(m,f)]] #only new technologies (not mature yet)
        self.MFT_NEW_YEARLY_FIRST_STAGE_MIN0 = [(m,f,t) for m in self.M_MODES for f in self.FM_FUEL[m] for t in self.T_YEARLY_TIME_FIRST_STAGE if (not self.tech_is_mature[(m,f)] and t!=self.T_YEARLY_TIME_FIRST_STAGE[0])]
        self.MFT_NEW_YEARLY_SECOND_STAGE = [(m,f,t) for m in self.M_MODES for f in self.FM_FUEL[m] for t in self.T_YEARLY_TIME_SECOND_STAGE if not self.tech_is_mature[(m,f)]]
        self.MFT_NEW_FIRST_PERIOD = [(m,f,t) for m in self.M_MODES for f in self.FM_FUEL[m] for t in [self.T_TIME_PERIODS[0]] if not self.tech_is_mature[(m,f)]]

        self.UT_UPG = [(e,f,t) for (e,f) in self.U_UPGRADE for t in self.T_TIME_PERIODS if (t <= self.T_TIME_PERIODS[-1] - self.LEAD_TIME_EDGE_UPG[(e,f)]) and (t in self.T_TIME_FIRST_STAGE) ]       
        self.UT_UPG_CONSTR = [(e,f,t) for (e,f) in self.U_UPGRADE for t in self.T_TIME_PERIODS_OPERATIONAL]  

        #
        #       WITH SCENARIOS
        #

        def combinations(list_of_tuples, list):
            list_of_tuples2 = []
            for tpl in list_of_tuples:
                for l in list:
                    tpl2 = tpl + (l,)
                    list_of_tuples2.append(tpl2) 
            return list_of_tuples2

        self.AFPT_S =          combinations(self.AFPT,self.S_SCENARIOS)
        self.APT_CONSTR_S =    combinations(self.APT_CONSTR,self.S_SCENARIOS)
        self.AFVT_S =          combinations(self.AFVT,self.S_SCENARIOS)
        self.AVT_CONSTR_S =    combinations(self.AVT_CONSTR,self.S_SCENARIOS)
        self.EAT_INV_CONSTR_S = combinations(self.EAT_INV_CONSTR,self.S_SCENARIOS)
        self.E_EDGES_INV_S = combinations(self.E_EDGES_INV,self.S_SCENARIOS)
        self.EFT_CHARGE_S = combinations(self.EFT_CHARGE,self.S_SCENARIOS)
        self.EFT_CHARGE_CONSTR_S = combinations(self.EFT_CHARGE_CONSTR,self.S_SCENARIOS)
        self.ET_INV_S = combinations(self.ET_INV,self.S_SCENARIOS)
        self.KPT_S = combinations(self.KPT,self.S_SCENARIOS)
        self.KVT_S = combinations(self.KVT,self.S_SCENARIOS)
        self.K_uni_VT_S = combinations(self.K_uni_VT,self.S_SCENARIOS)
        self.NM_CAP_S = combinations(self.NM_CAP,self.S_SCENARIOS)
        self.NM_CAP_T_CONSTR_S = combinations(self.NM_CAP_T_CONSTR,self.S_SCENARIOS)
        self.NM_CAP_T_S = combinations(self.NM_CAP_T,self.S_SCENARIOS)
        self.NM_CAP_INCR_T_S = combinations(self.NM_CAP_INCR_T,self.S_SCENARIOS)
        self.NM_CAP_INCR_S = combinations(self.NM_CAP_INCR,self.S_SCENARIOS)  
        self.MFT_S = combinations(self.MFT,self.S_SCENARIOS)
        self.MFT_MATURITY_CONSTR_S = combinations(self.MFT_MATURITY_CONSTR,self.S_SCENARIOS)
        self.MFT_NEW_YEARLY_S = combinations(self.MFT_NEW_YEARLY,self.S_SCENARIOS)
        self.MFT_NEW_S = combinations(self.MFT_NEW,self.S_SCENARIOS)
        self.MFT_MIN0_S = combinations(self.MFT_MIN0,self.S_SCENARIOS)
        self.MT_MIN0_S = combinations(self.MT_MIN0,self.S_SCENARIOS)
        self.MFT_INIT_TRANSP_SHARE_S = combinations(self.MFT_INIT_TRANSP_SHARE,self.S_SCENARIOS)
        self.MFT_NEW_FIRST_PERIOD_S = combinations(self.MFT_NEW_FIRST_PERIOD,self.S_SCENARIOS)
        self.MFT_NEW_YEARLY_FIRST_STAGE_MIN0_S = combinations(self.MFT_NEW_YEARLY_FIRST_STAGE_MIN0,self.S_SCENARIOS)
        self.MFT_NEW_YEARLY_SECOND_STAGE_S = combinations(self.MFT_NEW_YEARLY_SECOND_STAGE,self.S_SCENARIOS)
        self.MF_S = combinations(self.MF,self.S_SCENARIOS)
        self.MT_S = combinations(self.MT,self.S_SCENARIOS)
        self.MFT_CONSTR_S = combinations(self.MFT_CONSTR,self.S_SCENARIOS)
        self.M_MODES_S = combinations([(m,) for m in self.M_MODES],self.S_SCENARIOS)
        self.NMFVT_CONSTR_S = combinations(self.NMFVT_CONSTR,self.S_SCENARIOS)
        self.ODPTS_CONSTR_S = combinations(self.ODPTS_CONSTR,self.S_SCENARIOS)
        self.TS_S = combinations(self.TS,self.S_SCENARIOS)
        self.TS_NO_BASE_YEAR_S = combinations(self.TS_NO_BASE_YEAR,self.S_SCENARIOS)
        self.TS_CONSTR_S = combinations(self.TS_CONSTR,self.S_SCENARIOS)
        self.TS_NO_BASE_YEAR_CONSTR_S = combinations(self.TS_NO_BASE_YEAR_CONSTR,self.S_SCENARIOS)
        self.T_TIME_PERIODS_S = combinations([(t,) for t in self.T_TIME_PERIODS],self.S_SCENARIOS)
        self.UT_UPG_S = combinations(self.UT_UPG,self.S_SCENARIOS)
        self.UT_UPG_CONSTR_S = combinations(self.UT_UPG_CONSTR,self.S_SCENARIOS)


        #DERIVED PARAMETERS

        #find the "cheapest" product group per vehicle type. 
        self.cheapest_product_per_vehicle = {(m,f,t,v):None for m in self.M_MODES for f in self.FM_FUEL[m] for t in self.T_TIME_PERIODS for v in self.VEHICLE_TYPES_M[m]}
        for m in self.M_MODES: 
            for f in self.FM_FUEL[m]: 
                for t in self.T_TIME_PERIODS:
                    for v in self.VEHICLE_TYPES_M[m]:
                        cheapest_product = None
                        lowest_cost = 200000000
                        for p in self.PV_PRODUCTS[v]:
                            if self.C_TRANSP_COST_NORMALIZED[(m,f,p,t)] < lowest_cost:
                                lowest_cost = self.C_TRANSP_COST_NORMALIZED[(m,f,p,t)]
                                cheapest_product = p
                        self.cheapest_product_per_vehicle[(m,f,t,v)] = cheapest_product
        
        self.KFPT = []
        self.KFPT_S = []
        # for k,p,t in self.KPT:
        #     result = [i[2] for i in self.K_PATH_DICT[k]]
        #     modes = [] 
        #     [modes.append(x) for x in result if x not in modes] 
        #     modes = tuple(modes)
        #     if len(modes) == 1:
        #         modes = modes + modes
        #     for fuels in self.FM_MULTI[modes]:
        #         for s in self.S_SCENARIOS:
        #             self.KFPT_S.append((k,) + (fuels,) + (p,) + (t,) + (s,))
        #         self.KFPT.append((k,) + (fuels,) + (p,) + (t,))
        # For each old k,p,t,s combination, we combine with all the possible fuel combination of the modes of k
        # Ex: k = (("Hamar", "Trondheim", "Rail", 1), ("Trondheim", "Bodø", "Rail", 1), ("Bodø", "Narvik", "Road", 1))
        # the combination of fuels are self.FM_MULTI[("Rail", "Road")], with N elements
        # we have N new (k,fuels) combination
        self.KFVT = []
        self.K_uni_FVT = []
        self.KFVT_S = []
        self.K_uni_FVT_S = []
        
        if TIMES_data is not None:
            for t in self.T_TIME_PERIODS:
                for k_uni in self.UNI_MODAL_PATHS:
                    for fuels in self.FM_MULTI_K[k_uni]:
                        for v in self.V_VEHICLE_TYPES:
                            self.K_uni_FVT.append((k_uni,) + (fuels,) + (v,) + (t,))
                            self.KFVT.append((k_uni,) + (fuels,) + (v,) + (t,))
                            for s in self.S_SCENARIOS:
                                self.K_uni_FVT_S.append((k_uni,) + (fuels,) + (v,) + (t,) + (s,))
                                self.KFVT_S.append((k_uni,) + (fuels,) + (v,) + (t,) + (s,))
                        for p in self.P_PRODUCTS:
                            for s in self.S_SCENARIOS:
                                self.KFPT_S.append((k_uni,) + (fuels,) + (p,) + (t,) + (s,))
                            self.KFPT.append((k_uni,) + (fuels,) + (p,) + (t,))
                            
                for k_multi in self.MULTI_MODE_PATHS:
                    for fuels in self.FM_MULTI_K[k_multi]:
                        for v in self.V_VEHICLE_TYPES:
                            for s in self.S_SCENARIOS:
                                self.KFVT_S.append((k_multi,) + (fuels,) + (v,) + (t,) + (s,))
                            self.KFVT.append((k_multi,) + (fuels,) + (v,) + (t,))
                        for p in self.P_PRODUCTS:
                            for s in self.S_SCENARIOS:
                                self.KFPT_S.append((k_multi,) + (fuels,) + (p,) + (t,) + (s,))
                            self.KFPT.append((k_multi,) + (fuels,) + (p,) + (t,))


print("Finished reading sets and classes.")





